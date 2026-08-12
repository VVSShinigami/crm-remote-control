from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
from openpyxl.drawing.image import Image


class ExcelReport:
    def __init__(self, realized_ids: list[int], unrealized_ids: list[int]):
        self.realized_ids = realized_ids
        self.unrealized_ids = unrealized_ids

    def generate(self, output_path: str) -> bool:
        total = len(self.realized_ids) + len(self.unrealized_ids)
        if total == 0:
            return False

        chart_buf = self._create_pie_chart(total)

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по выполнению"

        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws['A1'] = "Обработанные"
        ws['B1'] = "Не обработанные"
        for cell in ws[1]:
            cell.font = bold_font
            cell.alignment = center_alignment
            cell.border = border

        max_rows = max(len(self.realized_ids), len(self.unrealized_ids))
        for i in range(max_rows):
            row_num = i + 2
            if i < len(self.realized_ids):
                ws[f'A{row_num}'] = self.realized_ids[i]
            if i < len(self.unrealized_ids):
                ws[f'B{row_num}'] = self.unrealized_ids[i]
            for col in ['A', 'B']:
                ws[f'{col}{row_num}'].border = border

        ws['E1'] = "СТАТИСТИКА:"
        ws['E1'].font = bold_font
        ws['E2'] = "Всего:"
        ws['F2'] = total
        ws['E3'] = "Успешно обработано:"
        ws['F3'] = len(self.realized_ids)
        ws['E4'] = "Не обработано:"
        ws['F4'] = len(self.unrealized_ids)
        ws['E5'] = "Процент успеха:"
        success_rate = (len(self.realized_ids) / total) * 100
        ws['F5'] = f"{success_rate:.1f}%"

        if chart_buf:
            img = Image(chart_buf)
            img.anchor = 'H2'
            ws.add_image(img)

        for column in ws.columns:
            col_max = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    col_max = max(col_max, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = col_max + 2

        try:
            Path(output_path).parent.mkdir(parents=True, exist_ok=True)
            wb.save(output_path)
            print(f"[ExcelReport] Сохранено: {output_path}")
            return True
        except Exception as e:
            print(f"[ExcelReport] ОШИБКА сохранения: {e}")
            return False

    def _create_pie_chart(self, total: int) -> io.BytesIO | None:
        if total == 0:
            return None

        labels = ["Обработанные", "Не обработанные"]
        sizes = [len(self.realized_ids), len(self.unrealized_ids)]
        colors = ["#2F8732", "#C21508"]

        fig, ax = plt.subplots(figsize=(4, 3))
        wedges, texts, autotexts = ax.pie(
            sizes,
            labels=labels,
            colors=colors,
            autopct='%1.1f%%',
            startangle=90,
            textprops={'fontsize': 12}
        )
        for autotext in autotexts:
            autotext.set_color('white')
            autotext.set_fontweight('bold')
            autotext.set_fontsize(10)

        ax.set_title(
            f'Статистика обработки\nВсего элементов: {total}',
            fontsize=14,
            fontweight='bold',
            pad=20
        )
        ax.axis('equal')
        ax.legend(
            wedges, labels,
            title="Категории",
            loc="center left",
            bbox_to_anchor=(1, 0, 0.5, 1)
        )

        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100, bbox_inches='tight')
        buf.seek(0)
        plt.close(fig)
        return buf