import requests #type:ignore
import os
import mimetypes
import openpyxl
from openpyxl import load_workbook
import csv


class FileParser:
    def read_file(self, file_path: str) -> list[int]:
        if not self._file_exists(file_path):
            return []

        file_type = self._definition_file_type(file_path)

        if file_type == "text/plain":
            return self._read_txt(file_path)
        elif file_type and ("csv" in file_type or "spreadsheet" in file_type or "excel" in file_type):
            return self._read_xlsx(file_path) if "excel" in file_type or "spreadsheet" in file_type else self._read_csv(file_path)

        return []

    def _file_exists(self, file_path: str) -> bool:
        return os.path.isfile(file_path)

    def _definition_file_type(self, file_path: str) -> str | None:
        mime_type, _ = mimetypes.guess_type(file_path)
        return mime_type

    def _read_txt(self, file_path: str) -> list[int]:
        ids = []
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line:
                    try:
                        ids.append(int(line))
                    except ValueError:
                        pass
        return ids

    def _read_csv(self, file_path: str) -> list[int]:
        import csv
        ids = []
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if row and row[0]:
                    try:
                        ids.append(int(row[0].strip()))
                    except (ValueError, TypeError):
                        pass
        return ids

    def _read_xlsx(self, file_path: str) -> list[int]:
        from openpyxl import load_workbook
        ids = []
        wb = load_workbook(file_path, read_only=True)
        sheet = wb.active
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] is not None:
                try:
                    ids.append(int(row[0]))
                except (ValueError, TypeError):
                    pass
        wb.close()
        return ids